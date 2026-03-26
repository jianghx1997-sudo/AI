from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

THIS_DIR = Path(__file__).resolve().parent
if str(THIS_DIR) not in sys.path:
    sys.path.insert(0, str(THIS_DIR))

from update_topic_library_links import refresh_topic_library_links


ROOT = Path(__file__).resolve().parents[1]
OUTPUTS_DIR = ROOT / "outputs"
CANONICAL_REVIEW_XLSX = "01_\u5ba1\u6838\u7a3f.xlsx"
SHEET_NAME = "Review"


SECTION_FILL = PatternFill("solid", fgColor="F2E5DA")
TITLE_FILL = PatternFill("solid", fgColor="8C4B2D")
TITLE_FONT = Font(name="Microsoft YaHei", size=12, bold=True, color="FFFFFF")
HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="222222")
NORMAL_FONT = Font(name="Microsoft YaHei", size=11, color="222222")
WRAP_TOP = Alignment(vertical="top", wrap_text=True)
WRAP_CENTER = Alignment(vertical="center", wrap_text=True)


SECTION_TOPIC_INFO = "\u4e13\u9898\u4fe1\u606f"
SECTION_TITLE = "\u6807\u9898"
SECTION_COVER = "\u5c01\u9762"
SECTION_COVER_POINTS = "\u5c01\u9762\u5217\u70b9"
SECTION_CARDS = "\u77e5\u8bc6\u5361"
SECTION_HASHTAGS = "#\u8bdd\u9898"
SECTION_CONFIRM = "\u786e\u8ba4\u8bf4\u660e"

LABEL_TOPIC = "\u4e13\u9898"
LABEL_STATUS = "\u72b6\u6001"
LABEL_NOTE = "\u8bf4\u660e"
LABEL_COVER_TITLE = "\u4e3b\u6807\u9898"
LABEL_COVER_SUBTITLE = "\u526f\u6807\u9898"

VALUE_PENDING = "\u5f85\u4eba\u5de5\u786e\u8ba4"
VALUE_NOTE = (
    "\u672c\u5730\u5ba1\u6838\u8868\u683c\u662f\u5ba1\u6838\u9636\u6bb5\u552f\u4e00 source of truth\u3002"
    "\u53ea\u6709\u5728\u7528\u6237\u660e\u786e\u786e\u8ba4\u540e\uff0c\u624d\u5141\u8bb8\u8fdb\u5165\u51fa\u56fe\u3002"
)
VALUE_CONFIRM = (
    "\u8bf7\u76f4\u63a5\u4fee\u6539\u8fd9\u4efd\u672c\u5730\u5ba1\u6838\u8868\u683c\u3002"
    "\u53ea\u6709\u5728\u660e\u786e\u8bf4\u201c\u6587\u7a3f\u65e0\u8bef\u201d\u6216\u201c\u53ef\u4ee5\u51fa\u56fe\u201d\u540e\uff0c"
    "\u624d\u5141\u8bb8\u8fdb\u5165\u56fe\u7247\u9636\u6bb5\u3002"
)

COVER_POINT_HEADERS = [
    "\u5e8f\u53f7",
    "mark_raw",
    "mark_render",
    "\u6587\u672c",
]
CARD_HEADERS = [
    "\u5361\u7247ID",
    "\u6807\u9898",
    "mark_raw",
    "mark_render",
    "\u6b63\u65871",
    "\u6b63\u65872",
    "\u6b63\u65873",
    "\u6b63\u65874",
]


def normalize_cover_points(points: list[Any]) -> list[dict[str, str]]:
    out: list[dict[str, str]] = []
    for point in points:
        if isinstance(point, dict):
            out.append(
                {
                    "text": str(point.get("text", "")).strip(),
                    "mark_raw": str(point.get("mark_raw", "")).strip(),
                    "mark_render": str(point.get("mark_render", "")).strip(),
                }
            )
        elif isinstance(point, str):
            text = point.strip()
            if text:
                out.append({"text": text, "mark_raw": "", "mark_render": text[:1]})
    return [item for item in out if item["text"]]


def normalize_cards(cards: list[Any]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for idx, card in enumerate(cards, start=1):
        if not isinstance(card, dict):
            continue
        paragraphs = [str(item).strip() for item in card.get("paragraphs", []) if str(item).strip()]
        out.append(
            {
                "id": str(card.get("id", f"{idx:02d}")).strip() or f"{idx:02d}",
                "title": str(card.get("title", "")).strip(),
                "mark_raw": str(card.get("mark_raw", "")).strip(),
                "mark_render": str(card.get("mark_render", "")).strip(),
                "paragraphs": paragraphs[:4],
            }
        )
    return out


def set_col_widths(ws) -> None:
    widths = {
        "A": 12,
        "B": 30,
        "C": 16,
        "D": 16,
        "E": 54,
        "F": 54,
        "G": 54,
        "H": 54,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def set_base_style(ws) -> None:
    set_col_widths(ws)
    for row in range(1, 80):
        ws.row_dimensions[row].height = 24
    for row in range(23, 29):
        ws.row_dimensions[row].height = 92


def section_label(ws, cell: str, text: str, fill=SECTION_FILL, font=HEADER_FONT) -> None:
    ws[cell] = text
    ws[cell].fill = fill
    ws[cell].font = font
    ws[cell].alignment = WRAP_CENTER


def write_review_xlsx(source: dict[str, Any]) -> Path:
    date_text = str(source.get("date", "")).strip()
    topic = str(source.get("topic", "")).strip()
    title = str(source.get("title", "")).strip()
    cover = source.get("cover", {}) or {}
    cover_title = str(cover.get("title", "")).strip()
    cover_subtitle = str(cover.get("subtitle", "")).strip()
    cover_points = normalize_cover_points(cover.get("points", []) or [])
    cards = normalize_cards(source.get("cards", []) or [])
    hashtags = [str(tag).strip() for tag in source.get("hashtags", []) if str(tag).strip()]

    if not date_text or not topic:
        raise ValueError("source must contain non-empty date and topic")

    topic_dir = OUTPUTS_DIR / f"{date_text}_{topic}"
    topic_dir.mkdir(parents=True, exist_ok=True)
    out_path = topic_dir / CANONICAL_REVIEW_XLSX

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    set_base_style(ws)

    section_label(ws, "A1", SECTION_TOPIC_INFO, fill=TITLE_FILL, font=TITLE_FONT)
    ws["A2"] = LABEL_TOPIC
    ws["B2"] = topic
    ws["A3"] = LABEL_STATUS
    ws["B3"] = VALUE_PENDING
    ws["A4"] = LABEL_NOTE
    ws["B4"] = VALUE_NOTE
    for cell in ("A2", "A3", "A4"):
        ws[cell].font = HEADER_FONT
        ws[cell].alignment = WRAP_CENTER
    for cell in ("B2", "B3", "B4"):
        ws[cell].font = NORMAL_FONT
        ws[cell].alignment = WRAP_TOP
    ws.row_dimensions[4].height = 42

    section_label(ws, "A6", SECTION_TITLE, fill=TITLE_FILL, font=TITLE_FONT)
    ws["B6"] = title
    ws["B6"].font = NORMAL_FONT
    ws["B6"].alignment = WRAP_TOP

    section_label(ws, "A8", SECTION_COVER, fill=TITLE_FILL, font=TITLE_FONT)
    ws["A9"] = LABEL_COVER_TITLE
    ws["B9"] = cover_title
    ws["A10"] = LABEL_COVER_SUBTITLE
    ws["B10"] = cover_subtitle
    for cell in ("A9", "A10"):
        ws[cell].font = HEADER_FONT
        ws[cell].alignment = WRAP_CENTER
    for cell in ("B9", "B10"):
        ws[cell].font = NORMAL_FONT
        ws[cell].alignment = WRAP_TOP

    section_label(ws, "A12", SECTION_COVER_POINTS, fill=SECTION_FILL)
    for idx, text in enumerate(COVER_POINT_HEADERS, start=1):
        cell = ws.cell(row=13, column=idx)
        cell.value = text
        cell.font = HEADER_FONT
        cell.fill = SECTION_FILL
        cell.alignment = WRAP_CENTER
    for idx, point in enumerate(cover_points[:6], start=1):
        row = 13 + idx
        ws[f"A{row}"] = f"{idx:02d}"
        ws[f"B{row}"] = point["mark_raw"]
        ws[f"C{row}"] = point["mark_render"]
        ws[f"D{row}"] = point["text"]
        for col in ("A", "B", "C", "D"):
            ws[f"{col}{row}"].font = NORMAL_FONT
            ws[f"{col}{row}"].alignment = WRAP_TOP

    section_label(ws, "A20", SECTION_CARDS, fill=TITLE_FILL, font=TITLE_FONT)
    for idx, text in enumerate(CARD_HEADERS, start=1):
        cell = ws.cell(row=22, column=idx)
        cell.value = text
        cell.font = HEADER_FONT
        cell.fill = SECTION_FILL
        cell.alignment = WRAP_CENTER
    for idx, card in enumerate(cards[:6], start=1):
        row = 22 + idx
        ws.cell(row=row, column=1, value=card["id"])
        ws.cell(row=row, column=2, value=card["title"])
        ws.cell(row=row, column=3, value=card["mark_raw"])
        ws.cell(row=row, column=4, value=card["mark_render"])
        for paragraph_index, paragraph in enumerate(card["paragraphs"][:4], start=5):
            ws.cell(row=row, column=paragraph_index, value=paragraph)
        for col in range(1, 9):
            ws.cell(row=row, column=col).font = NORMAL_FONT
            ws.cell(row=row, column=col).alignment = WRAP_TOP

    section_label(ws, "A30", SECTION_HASHTAGS, fill=TITLE_FILL, font=TITLE_FONT)
    for idx, tag in enumerate(hashtags[:8], start=31):
        ws[f"A{idx}"] = tag
        ws[f"A{idx}"].font = NORMAL_FONT
        ws[f"A{idx}"].alignment = WRAP_TOP

    section_label(ws, "A40", SECTION_CONFIRM, fill=TITLE_FILL, font=TITLE_FONT)
    ws["B40"] = VALUE_CONFIRM
    ws["B40"].font = NORMAL_FONT
    ws["B40"].alignment = WRAP_TOP
    ws.row_dimensions[40].height = 48

    wb.save(out_path)
    refresh_topic_library_links()
    return out_path


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("source_json", type=Path)
    args = parser.parse_args()

    source = json.loads(args.source_json.read_text(encoding="utf-8-sig"))
    out_path = write_review_xlsx(source)
    print(out_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
