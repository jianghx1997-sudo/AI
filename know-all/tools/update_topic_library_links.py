from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Font


ROOT = Path(__file__).resolve().parents[1]
OUTPUTS_DIR = ROOT / "outputs"
WORKBOOK_PATH = OUTPUTS_DIR / "topic_library_final.xlsx"
SHEET_NAME = "Topics"
TOPIC_HEADER = "\u5907\u9009\u8bdd\u9898"
DRAFT_NAME = "01_\u5ba1\u6838\u7a3f.xlsx"
HYPERLINK_DISPLAY_RE = re.compile(r'^=HYPERLINK\(".*?","(.*)"\)$', re.IGNORECASE)


def find_column_by_header(worksheet, header_name: str) -> int:
    for column in range(1, worksheet.max_column + 1):
        value = worksheet.cell(row=1, column=column).value
        if isinstance(value, str) and value.strip() == header_name:
            return column
    raise KeyError(f"Header not found: {header_name}")


def topic_name_from_folder(folder_name: str) -> str:
    if "_" in folder_name:
        _, topic = folder_name.split("_", 1)
        if topic:
            return topic
    return folder_name


def folder_score(folder: Path) -> Tuple[int, str]:
    has_pack = (folder / "02_approved-note-pack.json").exists()
    has_images = any(folder.glob("*-cover.png"))
    date_prefix = folder.name.split("_", 1)[0] if "_" in folder.name else ""
    return (2 if has_pack else 0) + (1 if has_images else 0), date_prefix


def collect_draft_docs(outputs_dir: Path) -> Dict[str, Path]:
    draft_docs: Dict[str, Path] = {}
    best_meta: Dict[str, Tuple[int, str]] = {}
    for child in outputs_dir.iterdir():
        if not child.is_dir():
            continue
        draft_doc = child / DRAFT_NAME
        if not draft_doc.exists():
            continue
        topic = topic_name_from_folder(child.name)
        score = folder_score(child)
        if topic not in best_meta or score > best_meta[topic]:
            draft_docs[topic] = draft_doc
            best_meta[topic] = score
    return draft_docs


def normalize_topic_name(cell_value) -> str:
    if cell_value is None:
        return ""
    text = str(cell_value).strip()
    match = HYPERLINK_DISPLAY_RE.match(text)
    if match:
        return match.group(1).strip()
    return text


def refresh_topic_library_links() -> int:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK_PATH}")

    workbook = load_workbook(WORKBOOK_PATH)
    if SHEET_NAME not in workbook.sheetnames:
        raise KeyError(f"Worksheet not found: {SHEET_NAME}")

    worksheet = workbook[SHEET_NAME]
    topic_column = find_column_by_header(worksheet, TOPIC_HEADER)
    draft_docs = collect_draft_docs(OUTPUTS_DIR)
    auto_filter_ref = worksheet.auto_filter.ref

    linked = 0
    cleared = 0

    for row in range(2, worksheet.max_row + 1):
        topic_cell = worksheet.cell(row=row, column=topic_column)
        topic_name = normalize_topic_name(topic_cell.value)
        if not topic_name:
            continue

        draft_doc = draft_docs.get(topic_name)
        if draft_doc is None:
            if topic_cell.hyperlink is not None:
                if topic_cell.data_type == "f":
                    topic_cell.value = topic_name
                topic_cell.hyperlink = None
                topic_cell.font = Font(name="Calibri", size=11, color="000000", underline=None)
                cleared += 1
            continue

        absolute_target = str(draft_doc)
        if topic_cell.data_type == "f":
            topic_cell.value = topic_name
        topic_cell.hyperlink = absolute_target
        topic_cell.style = "Hyperlink"
        linked += 1

    worksheet.auto_filter.ref = auto_filter_ref
    workbook.save(WORKBOOK_PATH)

    print(f"Workbook: {WORKBOOK_PATH}")
    print(f"Draft docs found: {len(draft_docs)}")
    print(f"Hyperlinks updated: {linked}")
    print(f"Hyperlinks cleared: {cleared}")
    return 0


def main() -> int:
    return refresh_topic_library_links()


if __name__ == "__main__":
    raise SystemExit(main())
