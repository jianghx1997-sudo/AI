from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import tools.render_core as renderer


SHEET_NAME = "Review"


def normalize(value) -> str:
    return "" if value is None else str(value).strip()


def parse_review_xlsx(xlsx_path: Path) -> dict:
    wb = load_workbook(xlsx_path)
    ws = wb[SHEET_NAME]

    topic = normalize(ws["B2"].value)
    title = normalize(ws["B6"].value)
    cover_title = normalize(ws["B9"].value)
    cover_subtitle = normalize(ws["B10"].value)

    cover_points = []
    for row in range(14, 20):
        text = normalize(ws[f"D{row}"].value)
        if not text:
            continue
        cover_points.append(
            {
                "text": text,
                "mark_raw": normalize(ws[f"B{row}"].value),
                "mark_render": normalize(ws[f"C{row}"].value)[:1],
            }
        )

    cards = []
    row = 23
    while True:
        card_id = normalize(ws.cell(row=row, column=1).value)
        if not card_id:
            break
        title_value = normalize(ws.cell(row=row, column=2).value)
        mark_raw = normalize(ws.cell(row=row, column=3).value)
        mark_render = normalize(ws.cell(row=row, column=4).value)
        paragraphs = []
        for col in range(5, 9):
            para = normalize(ws.cell(row=row, column=col).value)
            if para:
                paragraphs.append(para)
        cards.append(
            {
                "id": card_id,
                "title": title_value,
                "mark_raw": mark_raw,
                "mark_render": mark_render,
                "paragraphs": paragraphs,
            }
        )
        row += 1

    hashtags = []
    row = 31
    while row <= 38:
        tag = normalize(ws[f"A{row}"].value)
        if tag:
            hashtags.append(tag)
        row += 1

    return {
        "topic": topic,
        "title": title,
        "cover": {
            "title": cover_title,
            "subtitle": cover_subtitle,
            "points": cover_points,
        },
        "cards": cards,
        "hashtags": hashtags,
        "render_options": {
            "top_tags": False,
            "bottom_tips": False,
            "closing_page": False,
            "theme_cover_required": True,
        },
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("review_xlsx", type=Path)
    args = parser.parse_args()

    review_xlsx = args.review_xlsx
    pack = parse_review_xlsx(review_xlsx)
    out_dir = review_xlsx.parent
    topic = pack["topic"]

    approved_path = out_dir / "02_approved-note-pack.json"
    approved_path.write_text(json.dumps(pack, ensure_ascii=False, indent=2), encoding="utf-8")

    cover_path = out_dir / f"{topic}-cover.png"
    renderer.render_cover(pack, cover_path)

    card_paths = []
    for card in pack["cards"]:
        path = out_dir / f"{topic}-card-{card['id']}.png"
        renderer.render_card(pack, card, path)
        card_paths.append(path)

    preview_path = out_dir / f"{topic}-preview-grid.png"
    renderer.render_preview([cover_path, *card_paths], preview_path)

    print(approved_path)
    print(cover_path)
    for path in card_paths:
        print(path)
    print(preview_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
